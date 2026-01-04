"""
Excel Exporter Module for Smart Attendance System
Provides professional Excel export with charts and formatting
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ExcelExporter:
    """Professional Excel exporter with charts and formatting"""
    
    # Color scheme
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    PRESENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ABSENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: str = "Attendance"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_attendance(self, attendance_df: pd.DataFrame, subject: str, 
                          total_students: int = None) -> str:
        """Export attendance to Excel with charts"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Add title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Attendance Report - {subject}"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Add headers
        headers = ['Enrollment', 'Name', 'Date', 'Time', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Add data
        for row_idx, row in enumerate(attendance_df.itertuples(), 5):
            ws.cell(row=row_idx, column=1, value=row.Enrollment).border = self.BORDER
            ws.cell(row=row_idx, column=2, value=str(row.Name)).border = self.BORDER
            ws.cell(row=row_idx, column=3, value=row.Date).border = self.BORDER
            ws.cell(row=row_idx, column=4, value=row.Time).border = self.BORDER
            status_cell = ws.cell(row=row_idx, column=5, value="Present")
            status_cell.fill = self.PRESENT_FILL
            status_cell.border = self.BORDER
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # Add summary sheet with charts
        self._add_summary_sheet(wb, len(attendance_df), total_students or len(attendance_df))
        
        # Save file
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f"{self.output_dir}/{subject}_{timestamp}.xlsx"
        wb.save(filename)
        return filename
    
    def _add_summary_sheet(self, wb: Workbook, present: int, total: int):
        """Add summary sheet with pie chart"""
        ws = wb.create_sheet("Summary")
        absent = max(0, total - present)
        
        # Summary data
        ws['A1'] = "Attendance Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Status"
        ws['B3'] = "Count"
        ws['A4'] = "Present"
        ws['B4'] = present
        ws['A5'] = "Absent"
        ws['B5'] = absent
        ws['A6'] = "Total"
        ws['B6'] = total
        
        # Style
        for row in range(3, 7):
            ws.cell(row=row, column=1).border = self.BORDER
            ws.cell(row=row, column=2).border = self.BORDER
        
        ws['A3'].fill = self.HEADER_FILL
        ws['A3'].font = self.HEADER_FONT
        ws['B3'].fill = self.HEADER_FILL
        ws['B3'].font = self.HEADER_FONT
        
        # Pie Chart
        if total > 0:
            pie = PieChart()
            pie.title = "Attendance Distribution"
            labels = Reference(ws, min_col=1, min_row=4, max_row=5)
            data = Reference(ws, min_col=2, min_row=3, max_row=5)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.width = 12
            pie.height = 8
            ws.add_chart(pie, "D3")


def export_to_excel(attendance_df: pd.DataFrame, subject: str, 
                    total_students: int = None) -> str:
    """Convenience function for quick export"""
    exporter = ExcelExporter()
    return exporter.export_attendance(attendance_df, subject, total_students)
